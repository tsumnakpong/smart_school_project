from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from sms.models import Student
from .models import Course
from google import genai
from django.conf import settings
from django.http import JsonResponse,response
from django.views.decorators.csrf import csrf_exempt

from django.contrib.admin.views.decorators import staff_member_required
from .models import Lesson, AIChatLog  # อย่าลืม import AIChatLog ที่เราสร้างใหม่
from .models import  QuizQuestion  # เพิ่ม QuizScore เข้าไปตรงนี้
import openpyxl
from django.http import HttpResponse
import json
import random
from django.db.models import Avg, Max
from .models import  QuizScore,ParentProfile,StudentProfile
from .utils import render_to_pdf # นำเข้าฟังก์ชันที่เพิ่งสร้าง
from datetime import datetime
import requests
from django.utils import timezone
from .models import Attendance, LeaveRequest, Schedule
from .models import Notification
from datetime import timedelta
from django.db.models import Count, Q
from .utils import generate_tran_script
from django.http import FileResponse
from .utils import generate_student_qr_base64
import time
import hashlib
from django.db.models import Count, Q, Case, When, IntegerField
from datetime import datetime
import calendar
from django.db.models import Avg, Max, Count, Q, Sum, Case, When, IntegerField


# lms/views.py
import openpyxl
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill
from sms.models import Student, Subject, Classroom  # เพิ่ม Classroom ตรงนี้



@login_required
def scanner_check_in(request, schedule_id):
    schedule = get_object_or_404(Schedule, id=schedule_id)
    
    # ดึงรายชื่อนักเรียนทุกคนที่อยู่ในห้อง (Classroom) ของตารางเรียนนี้
    students = StudentProfile.objects.filter(
    classroom=schedule.classroom
    ).select_related('user').order_by('student_id')
    
    if request.method == "POST":
        student_id = request.POST.get('student_id')
        # รับค่า status จากปุ่ม Toggle (ถ้าเป็นการสแกน QR ปกติจะให้เป็น 'present')
        status = request.POST.get('status', 'present') 
        
        try:
            student_profile = StudentProfile.objects.get(student_id=student_id)
            
            # บันทึกหรืออัปเดตสถานะ (มาเรียน/ขาดเรียน)
            attendance, created = Attendance.objects.update_or_create(
                student=student_profile.user,
                schedule=schedule,
                date=timezone.now().date(),
                defaults={'status': status}
            )
            
            return JsonResponse({
                'status': 'success', 
                'name': student_profile.user.get_full_name(),
                'time': timezone.now().strftime('%H:%M:%S'),
                'current_status': status
            })
        except StudentProfile.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'ไม่พบข้อมูลนักเรียน'})

    # ดึงข้อมูลการเช็คชื่อของวันนี้มา เพื่อให้ปุ่ม Toggle แสดงสถานะล่าสุด
    present_list = Attendance.objects.filter(
        schedule=schedule, 
        date=timezone.now().date(),
        status='present'
    ).values_list('student__id', flat=True)

    context = {
        'schedule': schedule,
        'students': students,
        'present_list': present_list, # ส่ง ID ของเด็กที่เช็คชื่อแล้วไปที่ Template
    }
    return render(request, 'lms/teacher_scanner.html', context)



@login_required
def export_attendance_excel(request, classroom_id):
    if not request.user.is_staff:
        return HttpResponse("Unauthorized", status=401)

    classroom = get_object_or_404(Classroom, id=classroom_id)
    now = datetime.now()
    month = int(request.GET.get('month', now.month))
    year = int(request.GET.get('year', now.year))

    # 1. สร้าง Workbook และ Sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance_{month}_{year}"

    # 2. จัดรูปแบบหัวตาราง
    headers = ['ลำดับ', 'เลขประจำตัว', 'ชื่อ-นามสกุล', 'มาเรียน', 'สาย', 'ลา', 'ขาด', 'ร้อยละการเข้าเรียน']
    ws.append([f"รายงานการเข้าเรียนชั้น {classroom.name} ประจำเดือน {month}/{year}"])
    ws.merge_cells('A1:H1')
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.append(headers)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for cell in ws[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 3. ดึงข้อมูลนักเรียนและเขียนลง Excel
    students = classroom.student_profiles.all().select_related('user')
    for index, profile in enumerate(students, start=1):
        stats = Attendance.objects.filter(
            student=profile.user,
            date__month=month,
            date__year=year
        ).aggregate(
            present=Count(Case(When(status='present', then=1), output_field=IntegerField())),
            late=Count(Case(When(status='late', then=1), output_field=IntegerField())),
            absent=Count(Case(When(status='absent', then=1), output_field=IntegerField())),
            leave=Count(Case(When(status='leave', then=1), output_field=IntegerField())),
        )
        
        total_days = (stats['present'] or 0) + (stats['late'] or 0) + (stats['absent'] or 0) + (stats['leave'] or 0)
        rate = ((stats['present'] or 0) + (stats['late'] or 0)) / total_days * 100 if total_days > 0 else 0

        row = [
            index,
            profile.student_id,
            profile.user.get_full_name(),
            stats['present'] or 0,
            stats['late'] or 0,
            stats['leave'] or 0,
            stats['absent'] or 0,
            f"{rate:.2f}%"
        ]
        ws.append(row)

    # 4. ตั้งค่าความกว้างคอลัมน์อัตโนมัติ
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    # 5. ส่งไฟล์ออกไปที่ Browser
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=Attendance_{classroom.name}_{month}_{year}.xlsx'
    wb.save(response)
    return response

@login_required
def monthly_attendance_report(request, classroom_id):
    if not request.user.is_staff:
        return redirect('lms:dashboard')

    classroom = get_object_or_404(Classroom, id=classroom_id)
    
    # รับค่าเดือนและปีจาก URL (ถ้าไม่มีให้ใช้เดือนปัจจุบัน)
    now = datetime.now()
    month = int(request.GET.get('month', now.month))
    year = int(request.GET.get('year', now.year))
    
    # ดึงรายชื่อนักเรียนในห้อง
    students = classroom.student_profiles.all().select_related('user')
    
    report_data = []
    for profile in students:
        # สรุปสถิติในเดือนนั้นๆ ของนักเรียนแต่ละคน
        stats = Attendance.objects.filter(
            student=profile.user,
            date__month=month,
            date__year=year
        ).aggregate(
            present=Count(Case(When(status='present', then=1), output_field=IntegerField())),
            late=Count(Case(When(status='late', then=1), output_field=IntegerField())),
            absent=Count(Case(When(status='absent', then=1), output_field=IntegerField())),
            leave=Count(Case(When(status='leave', then=1), output_field=IntegerField())),
        )
        
        # คำนวณร้อยละการเข้าเรียน
        total_days = stats['present'] + stats['late'] + stats['absent'] + stats['leave']
        attendance_rate = (stats['present'] + stats['late']) / total_days * 100 if total_days > 0 else 0
        
        report_data.append({
            'student': profile,
            'stats': stats,
            'rate': round(attendance_rate, 2)
        })

    context = {
        'classroom': classroom,
        'report_data': report_data,
        'selected_month': month,
        'month_name': calendar.month_name[month],
        'years': range(now.year - 1, now.year + 2),
        'months': range(1, 13),
    }
    return render(request, 'lms/monthly_report.html', context)



@login_required
def student_confirm_attendance(request, schedule_id, token):
    # ตรวจสอบว่า Token ยังไม่หมดอายุ (เช็คย้อนหลังได้ 1 ช่วงเวลาเพื่อกันเน็ตช้า)
    current_ts = int(time.time() // 15)
    valid_tokens = [
        hashlib.sha256(f"{schedule_id}{current_ts}SECRET_KEY".encode()).hexdigest(),
        hashlib.sha256(f"{schedule_id}{current_ts-1}SECRET_KEY".encode()).hexdigest()
    ]

    if token not in valid_tokens:
        return render(request, 'lms/attendance_error.html', {'msg': 'QR Code หมดอายุแล้ว กรุณาสแกนใหม่'})

    schedule = get_object_or_404(Schedule, id=schedule_id)
    
    # บันทึกเช็คชื่อ
    Attendance.objects.update_or_create(
        student=request.user,
        schedule=schedule,
        date=timezone.now().date(),
        defaults={'status': 'present'}
    )

    return render(request, 'lms/attendance_success.html', {'schedule': schedule})

@login_required
def teacher_show_qr(request, schedule_id):
    schedule = get_object_or_404(Schedule, id=schedule_id)
    
    # สร้าง Token ลับที่เปลี่ยนตามนาที (ป้องกันการโกง)
    # เราใช้ schedule_id + secret_key + ช่วงเวลา 15 วินาที
    timestamp = int(time.time() // 15) 
    secret_token = hashlib.sha256(f"{schedule_id}{timestamp}settings.SECRET_KEY".encode()).hexdigest()
    
    # สร้าง URL สำหรับให้เด็กสแกน (เช่น /lms/scan/confirm/token_abc123/)
    qr_data = request.build_absolute_uri(f"/lms/attendance/confirm/{schedule_id}/{secret_token}/")
    
    # ใช้ฟังก์ชันเดิมสร้าง QR Base64
    qr_image = generate_student_qr_base64(qr_data)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'qr_image': qr_image})

    return render(request, 'lms/teacher_show_qr.html', {
        'schedule': schedule,
        'qr_image': qr_image
    })







@login_required
def student_dashboard(request):
    # ตรวจสอบว่าเป็นนักเรียนจริงๆ
    try:
        student_profile = request.user.studentprofile 
    except StudentProfile.DoesNotExist:
        return redirect('lms:teacher_dashboard')

    # สร้าง QR Code จากรหัสนักเรียน
    qr_code_image = generate_student_qr_base64(student_profile.student_id)

    context = {
        'student': student_profile,
        'qr_code': qr_code_image,
    }
    return render(request, 'lms/student_dashboard.html', context)

@login_required
def export_student_grade_pdf(request, student_id):
    student = get_object_or_404(StudentProfile, id=student_id)
    # ดึงเกรดทั้งหมดของนักเรียนคนนี้
    grades = TermGrade.objects.filter(student=student.user).select_related('subject')
    
    pdf_buffer = generate_tran_script(student, grades)
    
    filename = f"Transcript_{student.student_id}.pdf"
    return FileResponse(pdf_buffer, as_attachment=True, filename=filename)





def send_line_notification(token, message):
    url = 'https://notify-api.line.me/api/notify'
    headers = {'Authorization': f'Bearer {token}'}
    data = {'message': message}
    requests.post(url, headers=headers, data=data)

# ตัวอย่างการใช้ใน save_quiz_score
# message = f"📢 แจ้งคะแนนสอบ: {student_name}\nวิชา: {subject}\nได้คะแนน: {score}/{total}"
# send_line_notification(parent_token, message)

def generate_quiz(request, lesson_id):
    lesson = Lesson.objects.get(id=lesson_id)
    
    # 1. ลองหาข้อสอบในฐานข้อมูลก่อน
    db_questions = QuizQuestion.objects.filter(lesson=lesson)
    
    if db_questions.exists() and db_questions.count() >= 5:
        # ถ้ามีข้อสอบใน DB ครบ 5 ข้อ ให้สุ่มมาใช้เลย (ประหยัด Quota AI)
        selected_questions = random.sample(list(db_questions), 5)
        quiz_data = []
        for q in selected_questions:
            quiz_data.append({
                "question": q.question_text,
                "options": [q.option_1, q.option_2, q.option_3, q.option_4],
                "answer": q.correct_answer
            })
        return JsonResponse({'quiz': quiz_data, 'source': 'database'})

    # 2. ถ้าใน DB ไม่มี หรือมีไม่พอ ค่อยใช้ AI (Gemini) เจนให้
    try:
        client = genai.Client(api_key=settings.GEMINI_API_KEY)
        # ... โค้ดเรียก Gemini เดิมของคุณครู ...
        # (อย่าลืมใส่ Try-Except ดัก Error 429 ด้วยนะครับ)
        return JsonResponse({'quiz': json.loads(response.text), 'source': 'ai'})
        
    except Exception as e:
        return JsonResponse({'error': 'Quota เต็มและไม่มีข้อสอบสำรองในระบบ'}, status=429)

def generate_quiz(request, lesson_id):
    lesson = Lesson.objects.get(id=lesson_id)
    client = genai.Client(api_key=settings.GEMINI_API_KEY)
    try:
        prompt = f"""
        จากเนื้อหาบทเรียน: {lesson.content}
        จงสร้างข้อสอบปรนัย 5 ข้อ (4 ตัวเลือก) โดยส่งกลับมาเป็นรูปแบบ JSON เท่านั้น
        รูปแบบ JSON:
        [
        {{
            "question": "คำถาม...",
            "options": ["ก...", "ข...", "ค...", "ง..."],
            "answer": 0  // index ของข้อที่ถูก (0-3)
        }}
        ]
        """
        
        response = client.models.generate_content(
            model="gemini-2.0-flash",
            contents=prompt,
            config={'response_mime_type': 'application/json'} # บังคับให้ตอบเป็น JSON
        )
        
        return JsonResponse({'quiz': json.loads(response.text)})
    except Exception as e:
        if "429" in str(e):
            return JsonResponse({
                'error': 'คุณครู AI ขอพักเหนื่อยสักครู่ (Quota Full) อีก 1 นาทีลองกดใหม่นะจ๊ะ'
            }, status=429)
        return JsonResponse({'error': str(e)}, status=500)

@staff_member_required
def export_chat_logs_excel(request):
    # 1. สร้าง Workbook และ Sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Chat Logs"

    # 2. สร้างหัวตาราง (Header)
    columns = ['วันที่-เวลา', 'นักเรียน', 'บทเรียน', 'คำถามจากนักเรียน', 'คำตอบจาก AI']
    ws.append(columns)

    # 3. ดึงข้อมูลจากฐานข้อมูล
    logs = AIChatLog.objects.all().select_related('student', 'lesson').order_by('-timestamp')

    # 4. ใส่ข้อมูลลงใน Sheet
    for log in logs:
        # ปรับรูปแบบเวลาให้สวยงามก่อนลง Excel
        local_time = log.timestamp.strftime('%d/%m/%Y %H:%M')
        ws.append([
            local_time,
            log.student.username,
            log.lesson.title,
            log.question,
            log.answer
        ])

    # 5. ตั้งค่า Response ให้ Browser ดาวน์โหลดเป็นไฟล์
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="AI_Learning_Logs.xlsx"'
    
    wb.save(response)
    return response

def ai_tutor_chat(request, lesson_id):
    if request.method != "POST":
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        # 1. ตั้งค่า Client (รูปแบบใหม่)
        client = genai.Client(api_key=settings.GEMINI_API_KEY)
        model_id = "gemini-2.0-flash" # ใช้ชื่อรุ่นที่คุณมีได้เลย
        
        lesson = Lesson.objects.get(id=lesson_id)
        user_message = request.POST.get('message')

        # 2. จัดการประวัติจาก Session (เหมือนเดิม)
        history_key = f'chat_history_{lesson_id}'
        history = request.session.get(history_key, [])

        # 3. ส่งคำสั่ง (รูปแบบใหม่ของ SDK v2)
        # เราส่ง System Instruction ไปใน config
        response = client.models.generate_content(
            model=model_id,
            contents=user_message,
            config={
                'system_instruction': f"คุณคือครูสอนวิชา {lesson.course.subject.name_th} เนื้อหา: {lesson.content}",
                'history': history,
            }
        )

        # 4. บันทึกประวัติใหม่ลงใน Session
        history.append({"role": "user", "parts": [{"text": user_message}]})
        history.append({"role": "model", "parts": [{"text": response.text}]})
        request.session[history_key] = history[-10:]
        request.session.modified = True

        # 5. บันทึก Log ลงฐานข้อมูล
        AIChatLog.objects.create(
            student=request.user,
            lesson=lesson,
            question=user_message,
            answer=response.text
        )

        return JsonResponse({'reply': response.text})

    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return JsonResponse({'reply': "เกิดข้อผิดพลาดในการเชื่อมต่อ AI"}, status=500)
        # -------------------------------------------------------

        

    except Lesson.DoesNotExist:
        return JsonResponse({'reply': "ไม่พบบทเรียนนี้จ้ะ"}, status=404)
    except Exception as e:
        print(f"❌ พบข้อผิดพลาด: {str(e)}")
        return JsonResponse({'reply': "ระบบขัดข้องนิดหน่อยจ้ะ ลองใหม่นะ"}, status=500)



def lesson_detail(request, lesson_id):
    lesson = get_object_or_404(Lesson, pk=lesson_id)
    # ดึงบทเรียนทั้งหมดในคอร์สเดียวกันมาทำเป็นเมนูด้านข้าง
    course_lessons = Lesson.objects.filter(course=lesson.course).order_by('order')
    
    context = {
        'lesson': lesson,
        'course_lessons': course_lessons,
    }
    return render(request, 'lms/lesson_detail.html', context)

# lms/views.py

def clear_chat_history(request, lesson_id):
    history_key = f'chat_history_{lesson_id}'
    if history_key in request.session:
        del request.session[history_key]
        request.session.modified = True
    return JsonResponse({'status': 'cleared'})





@staff_member_required
def teacher_dashboard(request):
    # ดึงค่าห้องเรียนจาก URL เช่น ?classroom=6/1 (ถ้าไม่มีให้เป็น None)
    selected_class = request.GET.get('classroom')

    # 1. สรุปคะแนนเฉลี่ยแยกตาม "ทุกห้องเรียน" (เพื่อทำปุ่มเลือกและกราฟเปรียบเทียบ)
    class_overview = QuizScore.objects.values('classroom').annotate(
        avg_overall=Avg('score'),
        total_students=Count('student', distinct=True)
    ).order_by('classroom')

    # 2. ดูบทเรียนที่เด็กสงสัยมากที่สุด (แชทเยอะที่สุด)
    chat_query = AIChatLog.objects.all()
    if selected_class:
        # สมมติว่านักเรียนมีฟิลด์ classroom หรือใช้ความสัมพันธ์อื่นในการกรอง
        chat_query = chat_query.filter(student__profile__classroom=selected_class)
    
    top_lessons = chat_query.values('lesson__title').annotate(
        q_count=Count('id')
    ).order_by('-q_count')[:5]

    # 3. รายการคะแนนสอบรายบุคคล (กรองตามห้องเรียนที่เลือก)
    score_query = QuizScore.objects.all().select_related('student', 'lesson')
    if selected_class:
        score_query = score_query.filter(classroom=selected_class)
    
    student_scores = score_query.order_by('-timestamp')[:30]

    # 4. สถิติรายบทเรียน (กรองตามห้องเรียนที่เลือก)
    quiz_stats = score_query.values('lesson__title').annotate(
        avg_score=Avg('score'),
        max_score=Max('score'),
        student_count=Count('id')
    ).order_by('-avg_score')

    return render(request, 'lms/teacher_dashboard.html', {
        'class_overview': class_overview,
        'top_lessons': top_lessons,
        'quiz_stats': quiz_stats,
        'student_scores': student_scores,
        'selected_class': selected_class,
    })



@login_required
def save_quiz_score(request, lesson_id):
    if request.method == "POST":
        try:
            import json
            data = json.loads(request.body)
            score = data.get('score')
            total = data.get('total')
            
            lesson = Lesson.objects.get(id=lesson_id)
            student_user = request.user
            
            # 1. บันทึกคะแนนลงฐานข้อมูล
            QuizScore.objects.create(
                student=student_user,
                lesson=lesson,
                score=score,
                total_questions=total
            )

            # 2. สร้างการแจ้งเตือนสำหรับนักเรียน (ในระบบ)
            Notification.objects.create(
                user=student_user,
                title="ทำแบบทดสอบเสร็จสิ้น",
                message=f"คุณทำแบบทดสอบบทเรียน '{lesson.title}' ได้คะแนน {score}/{total}",
                notif_type='score'
            )

            # 3. ค้นหาผู้ปกครองและแจ้งเตือน (ในระบบ)
            # ดึงข้อมูลนักเรียนเพื่อหาผู้ปกครองที่ดูแล
            try:
                student_profile = student_user.lms_student_profile
                parents = ParentProfile.objects.filter(my_students=student_profile)
                
                for parent in parents:
                    Notification.objects.create(
                        user=parent.user, # แจ้งเตือนไปที่ User ของผู้ปกครอง
                        title=f"แจ้งผลคะแนน: {student_user.get_full_name()}",
                        message=f"บุตรหลานของท่านทำคะแนนวิชา {lesson.course.subject.name_th} ได้ {score}/{total} คะแนน",
                        notif_type='score'
                    )
                    # --- ถ้าครูจะใช้ Line Notify ให้ใส่โค้ดส่งตรงนี้ได้เลยครับ ---
                    # send_line_notification(parent.line_token, message)
                    
            except Exception as e:
                print(f"⚠️ ไม่สามารถส่งแจ้งเตือนให้ผู้ปกครองได้: {e}")

            return JsonResponse({'status': 'success'})

        except Lesson.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': 'ไม่พบบทเรียน'}, status=404)
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)




@login_required
def parent_dashboard(request):
    # 1. เข้าถึงโปรไฟล์ผู้ปกครอง
    try:
        parent = request.user.lms_parent_profile 
    except (ParentProfile.DoesNotExist, AttributeError):
        return HttpResponse("บัญชีนี้ไม่ได้ลงทะเบียนเป็นผู้ปกครองในระบบ LMS ครับ")

    # 2. ดึงข้อมูลนักเรียนในความดูแล (ใช้ my_students ตามที่แก้ใน models)
    # เพิ่ม prefetch_related เพื่อดึง scores และ chat_logs มาล่วงหน้า (ลด Query ใน Loop)
    my_students = parent.my_students.all().select_related('user')

    student_data = []
    for student in my_students:
        # 3. ดึงคะแนนสอบทั้งหมดของนักเรียนคนนี้
        all_scores = QuizScore.objects.filter(student=student.user).select_related('lesson').order_by('-timestamp')
        
        # 4. คำนวณค่าทางสถิติ
        avg_score = all_scores.aggregate(Avg('score'))['score__avg'] or 0
        chat_count = AIChatLog.objects.filter(student=student.user).count()
        
        # 5. ตรรกะ Early Warning (เปรียบเทียบครั้งล่าสุดกับค่าเฉลี่ย)
        warning_msg = None
        latest_score = None
        
        if all_scores.exists():
            latest_score = all_scores[0].score  # คะแนนครั้งล่าสุด
            
            # เงื่อนไขเตือน: ถ้ามีประวัติสอบมากกว่า 1 ครั้ง และคะแนนครั้งล่าสุดต่ำกว่าค่าเฉลี่ยเดิม
            if all_scores.count() > 1:
                # คำนวณค่าเฉลี่ยก่อนหน้า (ไม่รวมครั้งล่าสุด) เพื่อเปรียบเทียบพัฒนาการ
                previous_avg = all_scores[1:].aggregate(Avg('score'))['score__avg'] or 0
                
                if latest_score < previous_avg:
                    diff = round(previous_avg - latest_score, 2)
                    warning_msg = f"⚠️ คะแนนสอบล่าสุดลดลง {diff} คะแนน เมื่อเทียบกับค่าเฉลี่ยเดิม"

        student_data.append({
            'info': student,
            'scores': all_scores[:5],  # ส่งประวัติ 5 ครั้งล่าสุดไปแสดง
            'chat_count': chat_count,
            'avg_score': round(avg_score, 2),
            'latest_score': latest_score,
            'warning_msg': warning_msg,
        })

    return render(request, 'lms/parent_dashboard.html', {'student_data': student_data})




# lms/views.py

def export_pdf_report(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    scores = QuizScore.objects.filter(student=student.user).select_related('lesson__course__subject').order_by('-timestamp')
    
    # ดึงชื่อวิชาจากคะแนนสอบล่าสุด (ถ้ามี) 
    # หรือดึงจาก Enrollment วิชาแรกของนักเรียน
    subject_name = "วิชาวิทยาศาสตร์และเทคโนโลยี" # ค่าเริ่มต้น
    subject_code = "ว-----"
    if scores.exists():
        # ดึงจากความสัมพันธ์: QuizScore -> Lesson -> Course -> Subject
        subject_obj = scores[0].lesson.course.subject
        subject_name = subject_obj.name_th
        subject_code = subject_obj.code # ✅ สมมติว่าใน Model Subject มีฟิลด์ชื่อ code
    elif student.enrollments.exists():
        subject_obj = student.enrollments.first().subject
        subject_name = subject_obj.name_th
        subject_code = subject_obj.code

    context = {
        'student': student,
        'subject_code': subject_code, # ✅ ส่งรหัสวิชา
        'subject_name': subject_name, # ✅ ส่งชื่อวิชา
        'scores': scores,
        'avg_score': round(scores.aggregate(Avg('score'))['score__avg'] or 0, 2),
        'teacher_name': request.user.get_full_name() or request.user.username,
        'print_date': datetime.now().strftime("%d %B %Y %H:%M"),
    }
    
    return render_to_pdf(None, context)





@login_required
def student_dashboard(request):
    student_user = request.user
    
    # 1. ดึงข้อมูล Profile ของนักเรียน (ใช้ชื่อที่ตั้งไว้ใน Model เช่น student_profile)
    # ใช้ try-except เผื่อกรณี User นี้ยังไม่มี Profile ผูกไว้
    try:
        student = student_user.student_profile
    except AttributeError:
        # ถ้าไม่ใช่ Profile นักเรียน ให้ส่งกลับไปหน้าอื่น หรือแสดง Error
        return render(request, 'lms/error.html', {'message': 'ไม่พบข้อมูลนักเรียนสำหรับบัญชีนี้'})

    # 2. คำนวณสถิติคะแนน (MIS for Student)
    my_scores = QuizScore.objects.filter(student=student_user)
    avg_score = my_scores.aggregate(Avg('score'))['score__avg'] or 0
    total_quiz = my_scores.count()

    # 3. ดึงการแจ้งเตือน (Notifications)
    notifications = Notification.objects.filter(user=student_user).order_by('-created_at')[:5]
    unread_notif = Notification.objects.filter(user=student_user, is_read=False).count()

    # 4. รายวิชาที่ลงทะเบียน (Optimize ด้วย prefetch_related)
    enrollments = student.enrollments.all().select_related('subject', 'subject__course')
    
    course_list = []
    for enrollment in enrollments:
        # เข้าถึง Course ผ่าน Subject
        course = getattr(enrollment.subject, 'course', None)
        if course:
            # ดึงบทเรียนแรกสุดมาแสดงปุ่ม "เข้าสู่บทเรียน"
            first_lesson = course.lessons.order_by('order').first()
            course_list.append({
                'enrollment': enrollment,
                'course': course,
                'first_lesson': first_lesson,
            })

    # 5. รวบรวมข้อมูลส่งไปยัง Template
    context = {
        'student': student,
        'avg_score': round(avg_score, 2),
        'total_quiz': total_quiz,
        'notifications': notifications,
        'unread_notif': unread_notif,
        'course_list': course_list,
    }
    
    return render(request, 'lms/student_dashboard.html', context)

# lms/views.py

@login_required
def quiz_history(request):
    student_user = request.user
    
    # ดึงประวัติการสอบทั้งหมด เรียงจากล่าสุดขึ้นก่อน
    # select_related('lesson', 'lesson__course__subject') เพื่อลด Query
    history = QuizScore.objects.filter(student=student_user).select_related(
        'lesson__course__subject'
    ).order_by('-timestamp')

    # คำนวณภาพรวม
    total_attempts = history.count()
    highest_score = history.aggregate(Max('score'))['score__max'] or 0
    avg_score = history.aggregate(Avg('score'))['score__avg'] or 0

    context = {
        'history': history,
        'total_attempts': total_attempts,
        'highest_score': highest_score,
        'avg_score': round(avg_score, 2),
    }
    return render(request, 'lms/quiz_history.html', context)

# lms/views.py
from django.db.models import Avg, Count, Max
from .models import QuizScore, Lesson, AIChatLog # อย่าลืม import AIChatLog

@login_required
def teacher_analytics(request):
    # ตรวจสอบว่าเป็นครูหรือไม่ (สมมติว่าใช้ is_staff หรือกลุ่ม Teacher)
    if not request.user.is_staff:
        return redirect('lms:student_dashboard')

    # 1. ข้อมูลคะแนนเฉลี่ยแยกตามบทเรียน (เพื่อดูว่าบทไหนยากที่สุด)
    lesson_stats = QuizScore.objects.values('lesson__title').annotate(
        avg_score=Avg('score'),
        total_students=Count('student', distinct=True)
    ).order_by('lesson__order')

    # 2. ข้อมูลการใช้งาน AI Tutor (เด็กถามอะไร AI บ่อยที่สุด?)
    # เราจะดูว่าบทเรียนไหนที่มีการแชทถาม AI มากที่สุด
    ai_stats = AIChatLog.objects.values('lesson__title').annotate(
        chat_count=Count('id')
    ).order_by('-chat_count')[:5]

    # 3. ค้นหาเด็กกลุ่มเสี่ยง (คะแนนเฉลี่ย < 50%)
    at_risk_students = QuizScore.objects.values('student__first_name', 'student__last_name').annotate(
        my_avg=Avg('score')
    ).filter(my_avg__lt=5) # สมมติคะแนนเต็ม 10 ถ้าได้น้อยกว่า 5 คือกลุ่มเสี่ยง

    context = {
        'lesson_stats': lesson_stats,
        'ai_stats': ai_stats,
        'at_risk_students': at_risk_students,
    }
    return render(request, 'lms/teacher_analytics.html', context)

# lms/views.py

@login_required
def view_schedule(request):
    days = [1, 2, 3, 4, 5]
    day_names = ['จันทร์', 'อังคาร', 'พุธ', 'พฤหัสบดี', 'ศุกร์']
    
    schedules = Schedule.objects.filter(teacher=request.user).select_related('subject', 'classroom')
    
    # แก้ปัญหา NameError และใช้ชื่อฟิลด์ให้ตรงกับ Model
    schedule_by_day = {d: schedules.filter(day_of_week=d) for d in days}
    
    classroom_students = {}
    classrooms = schedules.values_list('classroom', flat=True).distinct()
    for class_id in classrooms:
        if class_id:
            # ใช้ classroom (ไม่ใช่ classroom_id) ตามที่ Model แจ้ง
            classroom_students[class_id] = StudentProfile.objects.filter(classroom=class_id).select_related('user')

    present_ids = Attendance.objects.filter(
        date=timezone.now().date(),
        status='present'
    ).values_list('student__id', flat=True)

    context = {
        'title': 'จัดการชั้นเรียนและตารางสอน',
        'days': zip(days, day_names),
        'schedule_by_day': schedule_by_day,
        'classroom_students': classroom_students,
        'present_ids': list(present_ids),
        'is_teacher': True,
    }
    return render(request, 'lms/schedule.html', context)


@login_required
def take_attendance(request, schedule_id):
    if not request.user.is_staff:
        return redirect('lms:dashboard')

    schedule = get_object_or_404(Schedule, id=schedule_id)
    classroom = schedule.classroom
    # ดึงรายชื่อนักเรียนในห้อง
    students = classroom.student_profiles.all().select_related('user')
    today = timezone.now().date()

    # 1. ดึงข้อมูลการแจ้งลาของวันนี้ (ที่อนุมัติแล้ว)
    # เพื่อเอาไปแสดงสถานะในหน้าเช็คชื่อ
    on_leave_students = LeaveRequest.objects.filter(
        start_date__lte=today,
        end_date__gte=today,
        status='approved'
    ).values('student_id', 'leave_type')
    
    # แปลงเป็น Dictionary {student_id: leave_type} เพื่อให้เช็คใน Template ง่ายๆ
    leave_dict = {item['student_id']: item['leave_type'] for item in on_leave_students}

    # 2. ดึงข้อมูลการเช็คชื่อที่มีอยู่แล้ว (กรณีครูเข้ามาแก้ไข)
    existing_attendance = Attendance.objects.filter(
        schedule=schedule, 
        date=today
    ).values('student_id', 'status')
    
    attendance_dict = {item['student_id']: item['status'] for item in existing_attendance}

    if request.method == "POST":
        for student_profile in students:
            user_id = student_profile.user.id
            status = request.POST.get(f'status_{user_id}')
            
            if status:
                Attendance.objects.update_or_create(
                    student=student_profile.user,
                    schedule=schedule,
                    date=today,
                    defaults={'status': status}
                )
        return redirect('lms:view_schedule')

    context = {
        'schedule': schedule,
        'students': students,
        'today': today,
        'leave_dict': leave_dict,        # ข้อมูลการลา
        'attendance_dict': attendance_dict, # ข้อมูลที่เช็คไปแล้ว
        'status_choices': Attendance.STATUS_CHOICES,
    }
    return render(request, 'lms/take_attendance.html', context)

@login_required
def submit_leave(request):
    if request.method == "POST":
        leave_type = request.POST.get('leave_type')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        reason = request.POST.get('reason')
        
        LeaveRequest.objects.create(
            student=request.user,
            leave_type=leave_type,
            start_date=start_date,
            end_date=end_date,
            reason=reason
        )
        # สร้าง Notification แจ้งเตือนครูประจำชั้น (ถ้ามีระบบครูประจำชั้น)
        return redirect('lms:student_dashboard')
        
    return render(request, 'lms/submit_leave.html')

@login_required
def teacher_leave_list(request):
    if not request.user.is_staff:
        return redirect('lms:dashboard')
    
    # ดูรายการลาที่รออนุมัติ
    pending_leaves = LeaveRequest.objects.filter(status='pending').order_by('-created_at')
    return render(request, 'lms/teacher_leave_list.html', {'pending_leaves': pending_leaves})



@login_required
def weekly_summary(request):
    if not request.user.is_staff:
        return redirect('lms:dashboard')

    today = timezone.now().date()
    last_week = today - timedelta(days=7)

    # 1. สถิติการมาเรียนรายวันในรอบสัปดาห์
    daily_attendance = Attendance.objects.filter(
        date__range=[last_week, today]
    ).values('date').annotate(
        present_count=Count('id', filter=Q(status='present')),
        absent_count=Count('id', filter=Q(status='absent')),
        late_count=Count('id', filter=Q(status='late')),
        leave_count=Count('id', filter=Q(status='leave'))
    ).order_by('date')

    # 2. รายชื่อนักเรียนที่ขาดเรียนบ่อย (ขาดเกิน 2 ครั้งในสัปดาห์นี้) - Early Warning
    frequent_absents = Attendance.objects.filter(
        date__range=[last_week, today],
        status='absent'
    ).values('student__first_name', 'student__last_name').annotate(
        absent_days=Count('id')
    ).filter(absent_days__gt=2).order_by('-absent_days')

    context = {
        'daily_attendance': daily_attendance,
        'frequent_absents': frequent_absents,
        'start_date': last_week,
        'end_date': today,
    }
    return render(request, 'lms/weekly_summary.html', context)

@login_required
def student_portfolio(request, student_id):
    # ดึงโปรไฟล์นักเรียน
    student = get_object_or_404(StudentProfile, id=student_id)
    
    # 1. ข้อมูลการเรียน (LMS Data)
    quiz_scores = QuizScore.objects.filter(student=student.user).select_related('lesson').order_by('-timestamp')
    avg_score = quiz_scores.aggregate(Avg('score'))['score__avg'] or 0
    
    # 2. ข้อมูลการมาเรียน (SMS Data)
    attendance = Attendance.objects.filter(student=student.user)
    attendance_stats = attendance.values('status').annotate(count=Count('status'))
    total_days = attendance.count()
    
    # 3. ข้อมูลการลา (Leave Data)
    leave_history = LeaveRequest.objects.filter(student=student.user).order_by('-start_date')

    context = {
        'student': student,
        'quiz_scores': quiz_scores,
        'avg_score': round(avg_score, 2),
        'attendance_stats': attendance_stats,
        'total_days': total_days,
        'leave_history': leave_history,
    }
    return render(request, 'lms/student_portfolio.html', context)

# lms/views.py

def calculate_grade(score):
    if score >= 80: return '4'
    elif score >= 75: return '3.5'
    elif score >= 70: return '3'
    elif score >= 65: return '2.5'
    elif score >= 60: return '2'
    elif score >= 55: return '1.5'
    elif score >= 50: return '1'
    else: return '0'

@login_required
def grade_summary(request, classroom_id):
    if not request.user.is_staff:
        return redirect('lms:dashboard')

    classroom = get_object_or_404(Classroom, id=classroom_id)
    students = classroom.student_profiles.all().select_related('user')
    
    # รวบรวมสรุปเกรดของนักเรียนทุกคนในห้อง
    grade_report = []
    for profile in students:
        # ดึงคะแนนเฉลี่ยจากทุกแบบทดสอบของนักเรียนคนนี้
        student_scores = QuizScore.objects.filter(student=profile.user)
        # สมมติเราคิดเป็น % จากคะแนนเต็มทั้งหมด
        total_earned = student_scores.aggregate(Sum('score'))['score__sum'] or 0
        total_possible = student_scores.aggregate(models.Sum('total_questions'))['total_questions__sum'] or 1
        
        final_percentage = (total_earned / total_possible) * 100
        grade = calculate_grade(final_percentage)
        
        grade_report.append({
            'student': profile,
            'final_percentage': round(final_percentage, 2),
            'grade': grade
        })

    context = {
        'classroom': classroom,
        'grade_report': grade_report,
    }
    return render(request, 'lms/grade_summary.html', context)

# lms/views.py


@login_required
def send_grade_notifications(request, classroom_id):
    if not request.user.is_staff:
        return JsonResponse({'status': 'error', 'message': 'Permission denied'}, status=403)

    classroom = get_object_or_404(Classroom, id=classroom_id)
    students = classroom.student_profiles.all().select_related('user')
    
    count = 0
    for profile in students:
        # คำนวณเกรด (ใช้ Logic เดียวกับหน้า Summary)
        student_scores = QuizScore.objects.filter(student=profile.user)
        total_earned = student_scores.aggregate(models.Sum('score'))['score__sum'] or 0
        total_possible = student_scores.aggregate(models.Sum('total_questions'))['total_questions__sum'] or 1
        final_percentage = (total_earned / total_possible) * 100
        grade = calculate_grade(final_percentage)

        # ค้นหาผู้ปกครองของเด็กคนนี้
        parents = ParentProfile.objects.filter(my_students=profile)
        
        for parent in parents:
            # 1. ส่ง Notification ในระบบ
            Notification.objects.create(
                user=parent.user,
                title=f"ประกาศผลการเรียน: {profile.user.get_full_name()}",
                message=f"ผลการเรียนเทอมล่าสุด วิชาฟิสิกส์และเทคโนโลยี ได้คะแนน {round(final_percentage, 2)}% เกรดที่ได้คือ: {grade}",
                notif_type='score'
            )
            
            # 2. (Optional) ส่ง Line Notify ถ้าครูมี Token ในระบบ
            # if parent.line_token:
            #     send_line_notification(parent.line_token, f"📢 ผลการเรียนของ {profile.user.first_name}: เกรด {grade}")
            
            count += 1

    messages.success(request, f"ส่งแจ้งเตือนผลการเรียนให้ผู้ปกครองเรียบร้อยแล้ว {count} ท่าน")
    return redirect('lms:grade_summary', classroom_id=classroom.id)