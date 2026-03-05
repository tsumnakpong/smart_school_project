# core/views.py

from django.contrib.auth.decorators import login_required
import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill
from lms.models import StudentProfile # หรือ path ที่ครูเก็บ Profile ไว้
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm # ✅ ตรวจสอบว่า import ตัวนี้มา
from django.contrib.auth.decorators import permission_required
import qrcode
import io
import base64





from django.contrib.auth import get_user_model

User = get_user_model()

@login_required
def profile_view(request):
    user = request.user
    
    # --- ส่วนที่ 1: จัดการการเปลี่ยนรหัสผ่าน (POST Method) ---
    p_form = PasswordChangeForm(user) # เตรียมฟอร์มเปล่าไว้สำหรับ GET
    
    if request.method == 'POST':
        # ใน HTML ของครูต้องมี <input type="hidden" name="p_form_submit" value="1"> ในฟอร์มเปลี่ยนรหัส
        if 'p_form_submit' in request.POST:
            p_form = PasswordChangeForm(user, request.POST)
            if p_form.is_valid():
                updated_user = p_form.save()
                update_session_auth_hash(request, updated_user) # ป้องกัน Session หลุด
                messages.success(request, 'เปลี่ยนรหัสผ่านสำเร็จแล้วครับคุณครู!')
                return redirect('profile')
            else:
                messages.error(request, 'ข้อมูลไม่ถูกต้อง กรุณาตรวจสอบข้อผิดพลาด')

    # --- ส่วนที่ 2: ดึงข้อมูลนักเรียน & สร้าง QR Code ---
    # ดึงรหัสนักเรียนจาก OneToOneField (related_name='student_profile')
    try:
        if hasattr(user, 'student_profile'):
            student_id = user.student_profile.student_id
            gpax = user.student_profile.gpax
        else:
            student_id = f"STAFF-{user.id:04d}"
            gpax = 0.0
    except Exception:
        student_id = "N/A"
        gpax = 0.0

    # สร้าง QR Code (ข้อมูลคือรหัสนักเรียน)
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_L,
        box_size=10,
        border=2,
    )
    qr.add_data(student_id)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="#1e293b", back_color="white")
    
    # แปลงรูปภาพเป็น Base64
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    # --- ส่วนที่ 3: ส่งข้อมูลทั้งหมดไปที่ Template ---
    context = {
        'p_form': p_form, # ฟอร์มเปลี่ยนรหัสผ่าน
        'student_id': student_id, # รหัสที่สกัดได้
        'qr_code': qr_base64, # รูป QR
        'gpax': gpax, # เกรดเฉลี่ย
    }
    
    return render(request, 'core/profile.html', context)

# --- ฟังก์ชันเสริม (Helper) สำหรับการ Import ข้อมูล ---
def validate_user_data(row_data):
    username, password, first_name, last_name, email, role, phone = row_data
    if not all([username, password, first_name, role]):
        return False, "ข้อมูลพื้นฐาน (Username, Password, ชื่อ, Role) ห้ามว่าง"
    
    valid_roles = [choice[0] for choice in User.ROLE_CHOICES]
    if str(role).lower() not in valid_roles:
        return False, f"Role '{role}' ไม่ถูกต้อง"
        
    return True, None


@permission_required('core.can_approve_repair', raise_exception=True)
def approve_repair_view(request):
    # เฉพาะคนที่มีสิทธิ์นี้เท่านั้นถึงจะเข้าฟังก์ชันนี้ได้
    pass




User = get_user_model()

def validate_user_data(row_data):
    username, password, first_name, last_name, email, role, phone = row_data
    if not all([username, password, first_name, role]):
        return False, "ข้อมูลพื้นฐาน (Username, Password, ชื่อ, Role) ห้ามว่าง"
    
    valid_roles = [choice[0] for choice in User.ROLE_CHOICES] # ดึงจาก Model User
    if str(role).lower() not in valid_roles:
        return False, f"Role '{role}' ไม่ถูกต้อง"
        
    return True, None

@staff_member_required
def import_users(request):
    if request.method == "POST" and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        report = {
            'success': 0, 'failed': 0,
            'roles': {'executive': 0, 'teacher': 0, 'student': 0, 'parent': 0}
        }
        errors = []
        seen_in_file = set()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            username = str(row[0]).strip() if row[0] else None
            
            # Validation
            is_valid, error_msg = validate_user_data(row)
            if not is_valid:
                errors.append(f"แถว {row_idx}: {error_msg}"); report['failed'] += 1; continue
            
            if username in seen_in_file or User.objects.filter(username=username).exists():
                errors.append(f"แถว {row_idx}: Username '{username}' ซ้ำ"); report['failed'] += 1; continue

            seen_in_file.add(username)
            password, first_name, last_name, email, role, phone = row[1:]

            try:
                user = User.objects.create_user(
                    username=username, password=str(password),
                    first_name=first_name, last_name=last_name,
                    email=email, role=role.lower(), phone=phone
                )
                
                # สร้าง Profile อัตโนมัติสำหรับนักเรียน
                if user.role == 'student':
                    StudentProfile.objects.get_or_create(user=user, student_id=username)
                
                report['success'] += 1
                report['roles'][user.role] += 1
            except Exception as e:
                errors.append(f"แถว {row_idx}: {str(e)}"); report['failed'] += 1

        return render(request, 'core/import_report.html', {'report': report, 'errors': errors})
    
    return render(request, 'core/import_users.html')

def download_user_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    headers = ['Username', 'Password', 'First Name', 'Last Name', 'Email', 'Role', 'Phone']
    ws.append(headers)
    
    # ตัวอย่างข้อมูล
    ws.append(['std69001', 'pass1234', 'สมชาย', 'ใจดี', 'somchai@mail.com', 'student', '0812345678'])
    
    # ตกแต่ง Header
    fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True); cell.fill = fill; cell.alignment = Alignment(horizontal='center')
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="User_Template.xlsx"'
    wb.save(response)
    return response


@login_required
def dashboard_redirect_view(request):
    role = request.user.role  # ดึงค่า role จาก Custom User Model ของคุณ

    if role == 'executive':
        return redirect('mis:executive_dashboard')  # ชื่อ name ใน urls.py ของแอป mis หรือ sms
    elif role == 'teacher':
        return redirect('lms:teacher_dashboard')    # ชื่อ name ใน urls.py ของแอป lms หรือ sms
    elif role == 'student':
        return redirect('lms:student_dashboard')    # ชื่อ name ใน urls.py ของแอป lms
    elif role == 'parent':
        return redirect('lms:parent_dashboard')     # ชื่อ name ใน urls.py
    else:
        return redirect('cms:home') # ถ้าไม่มี role ให้กลับไปหน้าแรกของ cms
    
